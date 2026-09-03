"""Download and normalize observed generator capital-cost source data."""

import argparse
import hashlib
from pathlib import Path

import openpyxl
import pandas as pd

from atb_config import load_config, resolve_atb_path
from scrape_atb_inputs import download_file


COLUMNS = [
    "technology",
    "technology_detail",
    "year",
    "metric",
    "value",
    "unit",
    "capacity_basis",
    "statistic",
    "geography",
    "dollar_year",
    "price_basis",
    "sample_count",
    "source_id",
    "source_file",
    "source_sheet",
    "source_table",
    "source_page_url",
    "source_data_url",
    "notes",
]

# Only the three mapped tables give national cost by technology; the rest split
# the same capacity by region, state, size, or panel type. Labels drift between
# editions ("Solar photovoltaic" -> "Solar PV" -> "Solar"), so every spelling stays.
EIA_TECHNOLOGY_MAP = {
    "Solar": ("upv", "Utility-scale solar (all reported solar)"),
    "Solar PV": ("upv", "Utility-scale solar PV"),
    "Solar photovoltaic": ("upv", "Utility-scale solar photovoltaic"),
    "Battery storage": ("battery", "Battery storage"),
    "Wind": ("wind-ons", "Wind (EIA broad energy-source category)"),
    "Natural gas": ("gas", "Natural gas (all reported technologies)"),
    "Petroleum liquids": ("petroleum", "Petroleum liquids"),
    "Biomass": ("biopower", "Biomass"),
    "Geothermal": ("geothermal", "Geothermal"),
    "Hydro": ("hydropower", "Hydroelectric"),
    "Hydroelectric": ("hydropower", "Hydroelectric"),
}

# Equipment rather than fuel; the only table with fuel cells. Unmapped on
# purpose: "Steam turbine" (ambiguous fuel), "... (as part of combined cycle)"
# (half a plant, whole in the table below), "Internal combustion engine" (gas
# and oil, no ReEDS counterpart).
EIA_PRIME_MOVER_MAP = {
    "Combustion turbine": ("gas", "Natural gas combustion turbine"),
    "Onshore wind turbine": ("wind-ons", "Onshore wind turbine"),
    "Photovoltaic": ("upv", "Photovoltaic"),
    "Energy storage, battery": ("battery", "Battery storage"),
    "Battery storage": ("battery", "Battery storage"),
    "Fuel cell": ("fuelcell", "Fuel cell"),
    "Geothermal turbines": ("geothermal", "Geothermal turbine"),
    "Hydroelectric turbine": ("hydropower", "Hydroelectric turbine"),
}

# Whole plants, so combined cycle arrives as one cost rather than split across
# its turbine halves. Closest observed match to ReEDS Gas-CC and Gas-CT.
EIA_GAS_TECHNOLOGY_MAP = {
    "Combined cycle": ("gas", "Natural gas combined cycle"),
    "Combustion turbine": ("gas", "Natural gas combustion turbine"),
    "Steam turbine": ("gas", "Natural gas steam turbine"),
    "Internal combustion engine": ("gas", "Natural gas internal combustion engine"),
}

# (title fragment, label map, source_table tag), matched against a lowercased
# table title in column A. Natural gas must precede prime mover so its title wins.
EIA_TABLES = (
    ("by major energy source", EIA_TECHNOLOGY_MAP, "major_energy_source"),
    ("natural gas generators installed", EIA_GAS_TECHNOLOGY_MAP, "natural_gas_technology"),
    ("by prime mover", EIA_PRIME_MOVER_MAP, "prime_mover"),
)


def _base_row(source_id, source, filename, sheet):
    return {
        "metric": "capital_cost",
        "source_id": source_id,
        "source_file": filename,
        "source_sheet": sheet,
        "source_page_url": source["page_url"],
        "source_data_url": source.get("data_url", ""),
    }


def _find_header_row(sheet, required_text):
    required = required_text.lower()
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if any(required in str(value).lower() for value in row if value is not None):
            return row_number
    raise ValueError(f"Could not find {required_text!r} in sheet {sheet.title!r}")


def extract_land_based_wind(path, source):
    """Extract LBNL's observed annual capacity-weighted installed wind cost."""
    sheet_name = "CapEx Over Time"
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    header = _find_header_row(sheet, "Commercial Operation Date")
    rows = []
    for year, value, sample_count in sheet.iter_rows(
        min_row=header + 1, max_col=3, values_only=True
    ):
        if isinstance(year, (int, float)) and isinstance(value, (int, float)):
            row = _base_row("land_based_wind", source, Path(path).name, sheet_name)
            row.update(
                technology="wind-ons",
                technology_detail="Land-based wind projects",
                year=int(year),
                value=float(value),
                unit="USD/kW",
                capacity_basis="nameplate",
                statistic="capacity_weighted_mean",
                geography="United States",
                dollar_year=2024,
                price_basis="real",
                sample_count=sample_count,
                notes="Observed project CapEx; 2024 COD values are preliminary.",
            )
            rows.append(row)
    workbook.close()
    return rows


def extract_utility_pv(path, source):
    """Extract LBNL's observed PV-only installed costs on AC and DC bases."""
    sheet_name = "CapEx Trend (PV-only)"
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    header = _find_header_row(sheet, "Solar COD")
    rows = []
    # Each tuple is (year, count, capacity-weighted mean) for one capacity basis.
    for basis, year_col, count_col, mean_col in (
        ("AC", 1, 2, 4),
        ("DC", 11, 12, 14),
    ):
        for values in sheet.iter_rows(min_row=header + 1, values_only=True):
            year = values[year_col - 1]
            value = values[mean_col - 1]
            if not isinstance(year, (int, float)) or not isinstance(value, (int, float)):
                continue
            row = _base_row("utility_pv", source, Path(path).name, sheet_name)
            row.update(
                technology="upv",
                technology_detail="Utility-scale PV-only projects",
                year=int(year),
                # Source values are $/W; convert to the common $/kW unit.
                value=float(value) * 1000,
                unit="USD/kW",
                capacity_basis=basis,
                statistic="capacity_weighted_mean",
                geography="United States",
                dollar_year=2024,
                price_basis="real",
                sample_count=values[count_col - 1],
                notes="Observed PV-only project CapEx; 2024 COD values are preliminary.",
            )
            rows.append(row)
    workbook.close()
    return rows


def extract_offshore_wind(path, source):
    """Extract completed-year offshore project CapEx series from NLR Figure 31."""
    sheet_name = "F31, Project CapEx"
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    header = _find_header_row(sheet, "Commercial Operation Date")
    last_year = int(source["last_historical_year"])
    series = (
        (3, "capacity_weighted_mean", "Global"),
        (5, "capacity_weighted_mean", "Europe and United States"),
        (7, "capacity_weighted_mean", "Asia"),
    )
    rows = []
    for values in sheet.iter_rows(min_row=header + 1, values_only=True):
        year = values[1]
        if not isinstance(year, (int, float)) or int(year) > last_year:
            continue
        for column, statistic, geography in series:
            value = values[column - 1]
            if not isinstance(value, (int, float)) or value <= 0:
                continue
            row = _base_row("offshore_wind", source, Path(path).name, sheet_name)
            row.update(
                technology="wind-ofs",
                technology_detail="Offshore wind projects",
                year=int(year),
                value=float(value),
                unit="USD/kW",
                capacity_basis="nameplate",
                statistic=statistic,
                geography=geography,
                dollar_year=2023,
                price_basis="real",
                sample_count="",
                notes=(
                    "Figure 31 annual project CapEx. The data file omits units; the "
                    "report's Figure 31 axis reads USD2023/kW and its section 1.2.2 "
                    "normalizes all costs to real 2023 USD (FX conversion, then U.S. "
                    "CPI). Post-2023 pipeline years are excluded."
                ),
            )
            rows.append(row)
    workbook.close()
    return rows


def _eia_table_for_title(title):
    """Return (label_map, table_tag) for a table title, or (None, None).

    The combined-cycle breakdown splits one plant across its turbine halves, so
    it is excluded even though its title matches the natural-gas fragment.
    """
    lowered = title.lower()
    if "at combined-cycle plants" in lowered:
        return None, None
    for fragment, label_map, tag in EIA_TABLES:
        if fragment in lowered:
            return label_map, tag
    return None, None


def extract_eia(path, source, year, data_url):
    """Extract every national cost-by-technology table in one EIA workbook."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = []
    label_map = None
    table_tag = None
    for values in sheet.iter_rows(values_only=True):
        label = values[0]
        value = values[1] if len(values) > 1 else None
        if isinstance(label, str) and "generators installed" in label.lower():
            label_map, table_tag = _eia_table_for_title(label)
            continue
        if label_map is None or not isinstance(label, str):
            continue
        entry = label_map.get(label.strip())
        if entry is None or not isinstance(value, (int, float)):
            continue
        technology, detail = entry
        row = _base_row("eia_generator_costs", source, Path(path).name, sheet.title)
        row["source_data_url"] = data_url
        row.update(
            technology=technology,
            technology_detail=detail,
            year=int(year),
            value=float(value),
            unit="USD/kW",
            capacity_basis="nameplate",
            statistic="capacity_weighted_mean",
            geography="United States",
            dollar_year=int(year),
            price_basis="nominal",
            sample_count="",
            source_table=table_tag,
            notes=(
                "EIA-860 generators installed in this year. Average construction "
                "cost is total cost divided by total capacity. Categories follow "
                "EIA definitions and are not one-to-one with ATB technologies."
            ),
        )
        rows.append(row)
    workbook.close()
    if not rows:
        raise ValueError(f"No EIA cost rows extracted from {path}")
    return rows

def _sha256(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _artifact(source_id, source, year=None):
    if source_id == "eia_generator_costs":
        filename = source["filename"].format(year=year)
        if year == int(source["last_year"]):
            data_url = source["current_data_url"]
        else:
            data_url = source["archive_data_url"].format(year=year)
    else:
        filename = source["filename"]
        data_url = source["data_url"]
    return filename, data_url


def scrape(config, selected="all", force=False, no_download=False):
    settings = config["historical_cost_sources"]
    output_dir = resolve_atb_path(settings["directory"])
    output_dir.mkdir(parents=True, exist_ok=True)
    allow_insecure = settings.get("allow_insecure_ssl_fallback", False)
    rows = []
    manifest = []
    source_items = settings["sources"].items()

    for source_id, source in source_items:
        if not source.get("enabled", True):
            continue
        selector = {
            "land_based_wind": "wind",
            "utility_pv": "solar",
            "offshore_wind": "offshore",
            "eia_generator_costs": "eia",
        }[source_id]
        if selected not in ("all", selector):
            continue
        years = (
            range(int(source["first_year"]), int(source["last_year"]) + 1)
            if source_id == "eia_generator_costs"
            else [None]
        )
        for year in years:
            filename, data_url = _artifact(source_id, source, year)
            path = output_dir / filename
            if not no_download:
                download_file(
                    data_url,
                    path,
                    force=force,
                    allow_insecure_ssl_fallback=allow_insecure,
                )
            elif not path.exists():
                raise FileNotFoundError(f"Missing local source file: {path}")

            manifest.append(
                {
                    "source_id": source_id,
                    "report_year": year or source.get("report_year", ""),
                    "page_url": source["page_url"],
                    "data_url": data_url,
                    "local_file": filename,
                    "size_bytes": path.stat().st_size,
                    "sha256": _sha256(path),
                }
            )
            if source_id == "land_based_wind":
                rows.extend(extract_land_based_wind(path, source))
            elif source_id == "utility_pv":
                rows.extend(extract_utility_pv(path, source))
            elif source_id == "offshore_wind":
                rows.extend(extract_offshore_wind(path, source))
            else:
                rows.extend(extract_eia(path, source, year, data_url))

    normalized = pd.DataFrame(rows, columns=COLUMNS).sort_values(
        ["technology", "source_id", "capacity_basis", "geography", "year"]
    )
    normalized_path = output_dir / settings["normalized_filename"]
    manifest_path = output_dir / settings["manifest_filename"]
    normalized.to_csv(normalized_path, index=False)
    pd.DataFrame(manifest).to_csv(manifest_path, index=False)
    print(f"\nNormalized {len(normalized):,} capital-cost observations:")
    print(f"  {normalized_path}")
    print(f"Recorded {len(manifest):,} source files and checksums:")
    print(f"  {manifest_path}")
    if not normalized.empty:
        summary = normalized.groupby(["source_id", "technology"])["year"].agg(
            ["min", "max", "count"]
        )
        print("\nCoverage")
        print(summary.to_string())


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Download official observed generator-cost workbooks and normalize "
            "their capital-cost series."
        )
    )
    parser.add_argument("--config", help="Path to config.yaml (default: ../config.yaml).")
    parser.add_argument(
        "--only",
        choices=["all", "wind", "solar", "offshore", "eia"],
        default="all",
        help="Process one source family (default: all).",
    )
    parser.add_argument(
        "--force", action="store_true", help="Replace source files already downloaded."
    )
    parser.add_argument(
        "--no-download",
        action="store_true",
        help="Rebuild normalized outputs from local source files only.",
    )
    args = parser.parse_args()
    scrape(
        load_config(args.config),
        selected=args.only,
        force=args.force,
        no_download=args.no_download,
    )


if __name__ == "__main__":
    main()
