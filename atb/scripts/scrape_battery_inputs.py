"""
Download the raw NLR ATB Excel workbook (used for battery capital costs).

The ATB summary flat file (ATBe.csv) only reports total battery CAPEX per
duration. ReEDS needs the split into power ($/kW) and energy ($/kWh) capital
cost components, which are only published in the ATB Excel workbook on the
"Utility-Scale Battery Storage" sheet.

This script downloads that raw workbook into scraped_input/. It does NOT write
an intermediate battery_costs_<year>.csv: generate_atb_files.py imports
``download_workbook`` / ``extract_battery_costs`` from this module and extracts
the battery costs directly from the raw workbook at runtime.

Usage (optional pre-fetch / preview):
    python scrape_battery_inputs.py --year 2024

Note: CSP cost ratios (csp_cost_ratios_*.csv) and historic capacity factors
(historic_capacity_factors.csv) are NOT produced here. Those are manually
maintained inputs (manual_input/) that do not come from the ATB workbook.
"""
import argparse
import os

import pandas as pd
import openpyxl

from atb_config import load_config, raw_file_path

THISDIR = os.path.dirname(os.path.abspath(__file__))
# parent atb/ directory; scraper outputs land in atb/scraped_input/.
ATBDIR = os.path.dirname(THISDIR)

BATTERY_SHEET = 'Utility-Scale Battery Storage'
# Block header label in the sheet -> output "cost" name expected by ReEDS.
BLOCKS = {
    'Capital Cost ($/kWh)': 'capcost_energy',   # energy component
    'Battery Power Capital Cost ($/kW)': 'capcost',  # power component
}
SCENARIOS = ['Advanced', 'Moderate', 'Conservative']


def download_workbook(year, cache_dir=None):
    """Compatibility helper that downloads the config-selected workbook."""
    from scrape_atb_inputs import download_file

    config = load_config()
    configured_year = config['atb']['year']
    if year != configured_year:
        raise ValueError(
            f'config.yaml selects ATB {configured_year}, not {year}. '
            'Update the config before downloading a different release.'
        )
    destination = raw_file_path(config, 'workbook')
    if cache_dir is not None:
        destination = os.path.join(cache_dir, config['raw_data']['workbook']['filename'])
    return str(download_file(config['raw_data']['workbook']['url'], destination))


def extract_battery_costs(xlsx_path):
    """Return a DataFrame of battery power/energy capital costs by scenario/year."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if BATTERY_SHEET not in wb.sheetnames:
        raise ValueError(f'Sheet {BATTERY_SHEET!r} not found in workbook.')
    rows = list(wb[BATTERY_SHEET].iter_rows(values_only=True))

    def find_label_row(label):
        for i, row in enumerate(rows):
            for v in row[:6]:
                if isinstance(v, str) and v.strip() == label:
                    return i
        raise ValueError(f'Label {label!r} not found in {BATTERY_SHEET}.')

    recs = []
    for label, costname in BLOCKS.items():
        r = find_label_row(label)
        # the row immediately below the label holds the year header
        yearhdr = rows[r + 1]
        ycols = [(j, int(v)) for j, v in enumerate(yearhdr)
                 if isinstance(v, (int, float)) and 2018 <= v <= 2061]
        if not ycols:
            raise ValueError(f'No year header found below {label!r}.')
        # the next three rows are Advanced / Moderate / Conservative
        for k, scen in enumerate(SCENARIOS):
            datarow = rows[r + 2 + k]
            labels = [str(v).strip() for v in datarow[:5] if isinstance(v, str)]
            if scen not in labels:
                raise ValueError(
                    f'Expected scenario {scen!r} below {label!r}, got {labels!r}.')
            rec = {'cost': costname, 'Scenario': scen}
            for j, y in ycols:
                rec[y] = datarow[j]
            recs.append(rec)

    df = pd.DataFrame(recs)
    yearcols = sorted(c for c in df.columns if isinstance(c, int))
    # order rows: capcost first, then capcost_energy; scenarios in canonical order
    cost_order = {'capcost': 0, 'capcost_energy': 1}
    scen_order = {s: i for i, s in enumerate(SCENARIOS)}
    df = (df.assign(_c=df['cost'].map(cost_order), _s=df['Scenario'].map(scen_order))
            .sort_values(['_c', '_s'])
            .drop(columns=['_c', '_s'])
            .reset_index(drop=True))
    return df[['cost', 'Scenario'] + yearcols]


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('--year', type=int, default=load_config()['atb']['year'],
                   help='ATB year; must match config.yaml.')
    p.add_argument('--outdir', default=None,
                   help='Optional override for the configured raw-data directory.')
    args = p.parse_args()

    xlsx = download_workbook(args.year, args.outdir)
    print(f'Raw workbook available at {xlsx}')
    # preview the extracted battery costs (not written to disk)
    df = extract_battery_costs(xlsx)
    print(f'Extracted battery costs: {df.shape[0]} rows x {df.shape[1]} cols')
    print(df.to_string(index=False, max_cols=6))


if __name__ == '__main__':
    main()
