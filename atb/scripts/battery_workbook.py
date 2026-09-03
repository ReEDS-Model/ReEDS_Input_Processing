"""Extract ReEDS battery cost components from a downloaded ATB workbook."""

from pathlib import Path

import openpyxl
import pandas as pd


BATTERY_SHEET = "Utility-Scale Battery Storage"
BLOCKS = {
    "Capital Cost ($/kWh)": "capcost_energy",
    "Battery Power Capital Cost ($/kW)": "capcost",
}
SCENARIOS = ["Advanced", "Moderate", "Conservative"]


def extract_battery_costs(xlsx_path):
    """Return battery power/energy capital costs by scenario and year."""
    workbook_path = Path(xlsx_path)
    workbook = openpyxl.load_workbook(
        workbook_path, read_only=True, data_only=True
    )
    try:
        if BATTERY_SHEET not in workbook.sheetnames:
            raise ValueError(
                f"Sheet {BATTERY_SHEET!r} not found in {workbook_path}."
            )
        rows = list(workbook[BATTERY_SHEET].iter_rows(values_only=True))
    finally:
        workbook.close()

    def find_label_row(label):
        for index, row in enumerate(rows):
            if any(
                isinstance(value, str) and value.strip() == label
                for value in row[:6]
            ):
                return index
        raise ValueError(f"Label {label!r} not found in {BATTERY_SHEET}.")

    records = []
    for label, cost_name in BLOCKS.items():
        label_row = find_label_row(label)
        year_header = rows[label_row + 1]
        year_columns = [
            (column, int(value))
            for column, value in enumerate(year_header)
            if isinstance(value, (int, float)) and 2018 <= value <= 2061
        ]
        if not year_columns:
            raise ValueError(f"No year header found below {label!r}.")

        for offset, scenario in enumerate(SCENARIOS):
            data_row = rows[label_row + 2 + offset]
            labels = [
                str(value).strip()
                for value in data_row[:5]
                if isinstance(value, str)
            ]
            if scenario not in labels:
                raise ValueError(
                    f"Expected scenario {scenario!r} below {label!r}, "
                    f"got {labels!r}."
                )
            record = {"cost": cost_name, "Scenario": scenario}
            record.update(
                {year: data_row[column] for column, year in year_columns}
            )
            records.append(record)

    frame = pd.DataFrame(records)
    year_columns = sorted(
        column for column in frame.columns if isinstance(column, int)
    )
    cost_order = {"capcost": 0, "capcost_energy": 1}
    scenario_order = {scenario: index for index, scenario in enumerate(SCENARIOS)}
    frame = (
        frame.assign(
            _cost_order=frame["cost"].map(cost_order),
            _scenario_order=frame["Scenario"].map(scenario_order),
        )
        .sort_values(["_cost_order", "_scenario_order"])
        .drop(columns=["_cost_order", "_scenario_order"])
        .reset_index(drop=True)
    )
    return frame[["cost", "Scenario", *year_columns]]
