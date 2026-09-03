"""Download and summarize the raw NLR ATB inputs configured in config.yaml."""

import argparse
from pathlib import Path

import openpyxl
import pandas as pd
import requests

from atb_config import load_config, raw_file_path


def _open_download(url, allow_insecure_ssl_fallback):
    """Open a verified download, optionally retrying after a TLS inspection error."""
    request_options = {
        "stream": True,
        "timeout": 300,
        "headers": {"User-Agent": "ReEDS-Input-Processing/1.0"},
    }
    try:
        return requests.get(url, **request_options)
    except requests.exceptions.SSLError as error:
        if not allow_insecure_ssl_fallback:
            raise RuntimeError(
                "TLS certificate verification failed. Install the required CA "
                "certificate in this environment or set "
                "raw_data.allow_insecure_ssl_fallback: true in config.yaml."
            ) from error

        import urllib3

        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        print("WARNING: TLS certificate verification failed.")
        print("         Retrying this public raw-data download with verify=False.")
        return requests.get(url, verify=False, **request_options)


def download_file(
    url, destination, force=False, allow_insecure_ssl_fallback=False
):
    """Download one file atomically, or reuse the existing local copy."""
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists() and not force:
        print(f"Using existing raw file: {destination}")
        return destination

    temporary = destination.with_suffix(destination.suffix + ".part")
    print(f"Downloading {url}")
    print(f"       into {destination}")
    try:
        with _open_download(url, allow_insecure_ssl_fallback) as response:
            response.raise_for_status()
            with temporary.open("wb") as stream:
                for chunk in response.iter_content(chunk_size=1024 * 1024):
                    if chunk:
                        stream.write(chunk)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    temporary.replace(destination)
    return destination


def summarize_flat_file(path):
    """Print a compact, user-visible overview of the downloaded ATBe CSV."""
    path = Path(path)
    preview = pd.read_csv(path, nrows=5, low_memory=False)
    header = list(preview.columns)
    technology_col = "technology" if "technology" in header else "Technology"
    year_col = "atb_year" if "atb_year" in header else None
    selected = [technology_col] + ([year_col] if year_col else [])
    summary = pd.read_csv(path, usecols=selected, low_memory=False)

    print("\nRaw flat file")
    print(f"  path: {path}")
    print(f"  size: {path.stat().st_size / 1024**2:,.1f} MiB")
    print(f"  rows: {len(summary):,}")
    if year_col:
        years = sorted(summary[year_col].dropna().unique().tolist())
        print(f"  ATB years: {years}")
    technologies = sorted(summary[technology_col].dropna().astype(str).unique())
    print(f"  technologies ({len(technologies)}): {', '.join(technologies)}")
    print("  first five rows:")
    print(preview.to_string(index=False, max_cols=8))


def summarize_workbook(path):
    """Print the path, size, and sheet names for the raw ATB workbook."""
    path = Path(path)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print("\nRaw workbook")
    print(f"  path: {path}")
    print(f"  size: {path.stat().st_size / 1024**2:,.1f} MiB")
    print(f"  sheets ({len(workbook.sheetnames)}): {', '.join(workbook.sheetnames)}")
    workbook.close()


def scrape(config, selected="all", force=False):
    """Download configured raw inputs and show what is now available."""
    raw = config["raw_data"]
    allow_insecure = raw.get("allow_insecure_ssl_fallback", False)
    if selected in ("all", "flat"):
        flat = download_file(
            raw["flat_file"]["url"],
            raw_file_path(config, "flat_file"),
            force=force,
            allow_insecure_ssl_fallback=allow_insecure,
        )
        summarize_flat_file(flat)
    if selected in ("all", "workbook"):
        workbook = download_file(
            raw["workbook"]["url"],
            raw_file_path(config, "workbook"),
            force=force,
            allow_insecure_ssl_fallback=allow_insecure,
        )
        summarize_workbook(workbook)


def main():
    parser = argparse.ArgumentParser(
        description="Download and summarize the raw NLR ATB flat file and workbook."
    )
    parser.add_argument("--config", help="Path to config.yaml (default: ../config.yaml).")
    parser.add_argument(
        "--only", choices=["all", "flat", "workbook"], default="all",
        help="Download only one raw input (default: all).",
    )
    parser.add_argument(
        "--force", action="store_true", help="Replace raw files that already exist."
    )
    args = parser.parse_args()
    scrape(load_config(args.config), selected=args.only, force=args.force)


if __name__ == "__main__":
    main()
